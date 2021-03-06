import sys ; sys.setrecursionlimit(sys.getrecursionlimit() * 5)
import re, csv, os, requests, time
from openpyxl import Workbook
from openpyxl.styles import Font
from youdao import *



# 本程序使用正则表达式提取文本中的单词，注意文本中一行只能匹配一个单词。

class WordTest:
    def __init__(self):
        root = os.path.dirname(__file__)
        for filename in os.listdir(root):
            if '.txt' in filename:
                print(filename)
                self.file_path = filename
                break;

    def translation(self, word):
        # 谷歌翻译api

        proxies = {'http': '223.242.225.46'}
        headers = {
            'user-agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.105 Safari/537.36'}
        url = 'http://translate.google.cn/translate_a/single?client=gtx&dt=t&dj=1&ie=UTF-8&sl=auto&tl=zh&q=' + word
        res = requests.get(url, headers=headers, timeout=10)
        return res.json()['sentences'][0]['trans']

    def detect(self):
        times = 0
        file_name = time.strftime('%Y-%m-%d', time.localtime(time.time()))
        file_path = self.file_path
        with open(file_path)as file:
            r_file = file.read()
        wb = Workbook()
        sheet = wb.active
        sheet.append(['translation', 'please input it here', 'outcome', 'reference','phonetic symbol'])
        for i in re.findall('[a-z].*?\s+', r_file):
            if 'adj.' not in i and 'vi.' not in i and 'vt.' not in i and 'n.' not in i and 'v.' not in i:
                times+=1
                word = i.strip()
                translation,phonetic = connect(word)
                #time.sleep(1)
                sheet.append([str(translation), '', f'=IF(B{times + 1}=D{times + 1},"答对了！","False")', word,str(phonetic)])
                print('写入成功！')
        for i in range(2, 22):
            color = Font(u'宋体', size=11, bold=True, italic=False, strike=False, color='000000')
            sheet.cell(row=i, column=3).font = color

        for i in range(2, 22):
            color = Font(u'宋体', size=11, bold=True, italic=False, strike=False, color='000000')
            sheet.cell(row=i, column=4).font = color

        print('一共检测到{0}个单词'.format(times))
        wb.save('../{}.xlsx'.format(file_name))


if __name__ == "__main__":
    WordTest().detect()
    os.remove('每日单词.txt')
